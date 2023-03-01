import openpyxl
import json

def read_excel_file(file_path, sheet_name):
    """Reads an excel file and returns a list of dictionaries"""
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Get the sheet
    sheet = workbook[sheet_name]
    # Get the header row
    header_row = sheet[1]
    # Get the data rows
    data_rows = sheet.iter_rows(min_row=2)
    # Create a list of dictionaries
    data = []
    for row in data_rows:
        row_data = {}
        for i in range(len(header_row)):
            row_data[header_row[i].value] = row[i].value
        data.append(row_data)
    return data

def create_new_excel_and_sheet(file_path, sheet_name):
    """Creates a new excel file and a new sheet"""
    workbook = openpyxl.Workbook()
    workbook.save(file_path)
    workbook = openpyxl.load_workbook(file_path)
    workbook.create_sheet(sheet_name)
    workbook.save(file_path)

def get_attribute_names(data):
    """Gets the attribute names from a json string"""
    attribute_names = []
    for item in data:   
        attribute_names.append(item)
    return attribute_names

def export_json_to_excel_file(file_path, sheet_name, data):
    """Exports a json string to an excel file"""
    # Create a new excel file and sheet
    create_new_excel_and_sheet(file_path, sheet_name)
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Get the sheet
    sheet = workbook[sheet_name]
    # Load the json data
    data = json.loads(data)

    attributes = get_attribute_names(data[0])

    j = 0
    for attribute in attributes:
        sheet.cell(row=1, column=j+1).value = attribute
        j=j+1

    # Write the data to the excel file
    for i in range(len(data)):
        attributes = get_attribute_names(data[i])
        j = 0
        for attribute in attributes:
            sheet.cell(row=i+2, column=j+1).value = data[i][attribute] 
            j=j+1
    # Save the workbook
    workbook.save(file_path)
   

if __name__ == "__main__":
    data = read_excel_file("Namelist.xlsx", "Name")
    print(data)
    # export_json_to_excel_file("Name.xlsx", "Key", json.dumps(data))